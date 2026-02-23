import os, io, re, uuid, traceback
import numpy as np
import cv2
import pytesseract
import pandas as pd
from PIL import Image
from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment,
                              Border, Side, numbers as xl_numbers)
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
ALLOWED = {'png', 'jpg', 'jpeg', 'bmp', 'tiff', 'webp', 'gif'}

def allowed(f): return '.' in f and f.rsplit('.',1)[1].lower() in ALLOWED


# ─────────────────────────────────────────────────────────────────
# IMAGE PREPROCESSING
# ─────────────────────────────────────────────────────────────────

def preprocess(img_bgr: np.ndarray) -> np.ndarray:
    """
    Full preprocessing pipeline:
    grayscale → denoise → CLAHE → adaptive threshold → deskew
    Returns binarised image ready for Tesseract.
    """
    h, w = img_bgr.shape[:2]
    # Upscale small images for better OCR
    if w < 1200:
        scale = 1600 / w
        img_bgr = cv2.resize(img_bgr,
                             (int(w*scale), int(h*scale)),
                             interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)

    # Shadow removal
    dilated = cv2.dilate(gray, cv2.getStructuringElement(cv2.MORPH_RECT,(15,15)))
    bg      = cv2.medianBlur(dilated, 21)
    diff    = 255 - cv2.absdiff(gray, bg)
    gray    = cv2.normalize(diff, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)

    # CLAHE contrast enhancement
    gray = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8)).apply(gray)

    # Bilateral filter (edge-preserving denoise)
    gray = cv2.bilateralFilter(gray, 9, 75, 75)

    # Sharpen
    k    = np.array([[-1,-1,-1],[-1,9,-1],[-1,-1,-1]])
    gray = np.clip(cv2.filter2D(gray,-1,k), 0, 255).astype(np.uint8)

    # Adaptive threshold
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 31, 15
    )

    # Morphological clean
    binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN,
                               cv2.getStructuringElement(cv2.MORPH_RECT,(2,2)))
    return binary


def deskew(binary: np.ndarray) -> np.ndarray:
    coords = np.column_stack(np.where(binary < 128))
    if len(coords) < 50: return binary
    angle = cv2.minAreaRect(coords)[-1]
    if angle < -45: angle = 90 + angle
    if abs(angle) < 0.5: return binary
    h, w  = binary.shape
    M     = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
    return cv2.warpAffine(binary, M, (w,h),
                          flags=cv2.INTER_CUBIC,
                          borderMode=cv2.BORDER_REPLICATE)


def detect_table_lines(binary: np.ndarray):
    """Detect horizontal and vertical ruling lines via morphology."""
    h, w  = binary.shape
    inv   = cv2.bitwise_not(binary)

    # Horizontal lines
    kh = cv2.getStructuringElement(cv2.MORPH_RECT, (max(w//5,30), 1))
    horiz = cv2.morphologyEx(inv, cv2.MORPH_OPEN, kh, iterations=2)

    # Vertical lines
    kv = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(h//5,30)))
    vert  = cv2.morphologyEx(inv, cv2.MORPH_OPEN, kv, iterations=2)

    grid  = cv2.add(horiz, vert)
    return horiz, vert, grid


# ─────────────────────────────────────────────────────────────────
# OCR EXTRACTION
# ─────────────────────────────────────────────────────────────────

def _clean(text: str) -> str:
    text = re.sub(r'\s+', ' ', text).strip()
    return ''.join(c for c in text if c.isprintable())


def ocr_with_data(pil_img: Image.Image, psm: int = 6):
    """
    Run Tesseract image_to_data and return list of word dicts:
    {text, conf, left, top, width, height}
    """
    cfg  = f'--oem 3 --psm {psm}'
    data = pytesseract.image_to_data(
        pil_img, config=cfg, output_type=pytesseract.Output.DICT
    )
    words = []
    for i, txt in enumerate(data['text']):
        txt  = _clean(str(txt))
        conf = int(data['conf'][i])
        if not txt or conf < 0:
            continue
        words.append({
            'text': txt,
            'conf': conf,
            'left': int(data['left'][i]),
            'top':  int(data['top'][i]),
            'w':    int(data['width'][i]),
            'h':    int(data['height'][i]),
        })
    return words


def cluster_rows(words: list, row_gap_ratio: float = 0.6) -> list:
    """Group words into rows by vertical proximity."""
    if not words: return []
    words = sorted(words, key=lambda w: (w['top'], w['left']))
    med_h = float(np.median([w['h'] for w in words])) or 14
    gap   = max(med_h * row_gap_ratio, 6)

    rows, cur = [], [words[0]]
    for word in words[1:]:
        cy_ref = np.mean([w['top'] + w['h']/2 for w in cur])
        cy_new = word['top'] + word['h'] / 2
        if abs(cy_new - cy_ref) <= gap:
            cur.append(word)
        else:
            rows.append(sorted(cur, key=lambda w: w['left']))
            cur = [word]
    rows.append(sorted(cur, key=lambda w: w['left']))
    return rows


def assign_columns(rows: list, img_width: int) -> list:
    """
    Detect column boundaries using X-axis occupancy projection,
    then assign each word to a column slot.
    Returns rows where each element is a list of word dicts per column.
    """
    if not rows: return []

    # Build occupancy mask
    occ = np.zeros(img_width, dtype=np.int32)
    for row in rows:
        for w in row:
            l = max(0, w['left'])
            r = min(img_width-1, w['left']+w['w'])
            if l < r: occ[l:r] += 1

    # Find gaps ≥ 10px
    in_gap, gs, gaps = False, 0, []
    for x in range(img_width):
        if occ[x] == 0:
            if not in_gap: in_gap, gs = True, x
        else:
            if in_gap and x-gs >= 10:
                gaps.append((gs, x))
            in_gap = False

    # Build column boundaries from gaps
    bounds, prev = [], 0
    for g_start, g_end in gaps:
        mid = (g_start + g_end) // 2
        bounds.append((prev, mid))
        prev = mid
    bounds.append((prev, img_width))

    if len(bounds) < 2:
        # Single column — each word is its own "column"
        return [[row] for row in rows]

    # Assign words to column slots
    result = []
    for row in rows:
        slots = [[] for _ in bounds]
        for word in row:
            cx  = word['left'] + word['w'] // 2
            idx = len(bounds) - 1
            for ci, (bl, br) in enumerate(bounds):
                if bl <= cx < br:
                    idx = ci; break
            slots[idx].append(word)
        result.append(slots)
    return result


def merge_slot(words: list) -> dict:
    """Merge multiple words in one cell slot into a single cell dict."""
    if not words:
        return {'text': '', 'conf': -1,
                'bbox': None}
    text = ' '.join(w['text'] for w in words)
    conf = int(np.mean([w['conf'] for w in words]))
    l = min(w['left']       for w in words)
    t = min(w['top']        for w in words)
    r = max(w['left']+w['w'] for w in words)
    b = max(w['top'] +w['h'] for w in words)
    return {
        'text': text,
        'conf': conf,
        'bbox': [l, t, r-l, b-t],   # [x, y, w, h]
    }


# ─────────────────────────────────────────────────────────────────
# COLUMN INTELLIGENCE
# ─────────────────────────────────────────────────────────────────

_RE_NUM = re.compile(r'^[\-\+]?[\d,]+(\.\d+)?$')

def _to_float(v):
    try:    return float(re.sub(r'[^\d.\-\+]', '', v))
    except: return None

def detect_numeric_cols(table: list, header_row: int = 0) -> list:
    """
    Returns list of column indices that are predominantly numeric
    (≥60% of data cells parse as numbers).
    """
    if not table or len(table) < 2: return []
    n_cols   = len(table[0])
    data     = table[header_row+1:]
    num_cols = []
    for ci in range(n_cols):
        vals  = [row[ci] for row in data if ci < len(row) and row[ci].strip()]
        if not vals: continue
        count = sum(1 for v in vals if _to_float(v) is not None)
        if count / len(vals) >= 0.60:
            num_cols.append(ci)
    return num_cols


def compute_col_sums(table: list, num_cols: list, header_row: int = 0) -> dict:
    """Compute column sums for numeric columns."""
    data   = table[header_row+1:]
    totals = {}
    for ci in num_cols:
        vals = [_to_float(row[ci]) for row in data
                if ci < len(row) and _to_float(row[ci]) is not None]
        if vals: totals[ci] = round(sum(vals), 4)
    return totals


# ─────────────────────────────────────────────────────────────────
# MAIN OCR PIPELINE
# ─────────────────────────────────────────────────────────────────

def run_ocr(image_path: str) -> dict:
    """
    Full pipeline. Returns structured dict:
    {
      data:        2-D list of strings,
      confidence:  2-D list of ints (0-100, -1 = empty),
      coordinates: 2-D list of [x,y,w,h] or null,
      accuracy:    float (overall average confidence),
      numeric_cols: list of col indices,
      col_sums:    {col_idx: sum},
      img_size:    [w, h]  (of the preprocessed image, for coord scaling)
    }
    """
    img_bgr = cv2.imread(image_path)
    if img_bgr is None:
        raise ValueError(f"Cannot read image: {image_path}")

    # Store original dimensions for coordinate mapping
    orig_h, orig_w = img_bgr.shape[:2]

    # Preprocess
    binary = preprocess(img_bgr)
    binary = deskew(binary)
    proc_h, proc_w = binary.shape

    # Scale factor from original to processed coords
    sx = proc_w / orig_w
    sy = proc_h / orig_h

    pil_img = Image.fromarray(binary)

    # Try multiple PSM modes, pick best result
    best_words, best_score = [], -1
    for psm in [6, 4, 11]:
        try:
            words = ocr_with_data(pil_img, psm=psm)
            score = sum(w['conf'] for w in words) / max(len(words), 1)
            if score > best_score and len(words) > best_score:
                best_words, best_score = words, score
        except Exception:
            continue

    if not best_words:
        return _empty_result()

    # Cluster into rows
    row_clusters = cluster_rows(best_words)

    # Assign to column slots
    col_slots = assign_columns(row_clusters, proc_w)

    # Build 2-D tables
    data_table  = []
    conf_table  = []
    coord_table = []

    for slot_row in col_slots:
        row_data, row_conf, row_coord = [], [], []
        for slot in slot_row:
            cell = merge_slot(slot)
            row_data.append(cell['text'])
            row_conf.append(cell['conf'])
            # Scale coords back to original image space
            if cell['bbox']:
                x, y, w, h = cell['bbox']
                scaled = [
                    int(x / sx), int(y / sy),
                    int(w / sx), int(h / sy)
                ]
                row_coord.append(scaled)
            else:
                row_coord.append(None)
        if any(t.strip() for t in row_data):
            data_table.append(row_data)
            conf_table.append(row_conf)
            coord_table.append(row_coord)

    if not data_table:
        return _empty_result()

    # Normalise row lengths
    max_cols = max(len(r) for r in data_table)
    data_table  = [r + ['']    * (max_cols - len(r)) for r in data_table]
    conf_table  = [r + [-1]   * (max_cols - len(r)) for r in conf_table]
    coord_table = [r + [None] * (max_cols - len(r)) for r in coord_table]

    # Column intelligence
    num_cols = detect_numeric_cols(data_table)
    col_sums = compute_col_sums(data_table, num_cols)

    # Overall accuracy
    all_confs = [c for row in conf_table for c in row if c >= 0]
    accuracy  = round(np.mean(all_confs), 1) if all_confs else 0.0

    return {
        'data':        data_table,
        'confidence':  conf_table,
        'coordinates': coord_table,
        'accuracy':    accuracy,
        'numeric_cols': num_cols,
        'col_sums':    {str(k): v for k, v in col_sums.items()},
        'img_size':    [orig_w, orig_h],
    }


def _empty_result():
    return {
        'data': [], 'confidence': [], 'coordinates': [],
        'accuracy': 0, 'numeric_cols': [], 'col_sums': {},
        'img_size': [0, 0]
    }


# ─────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────

def build_excel(table_data: list,
                conf_grid:  list,
                col_types:  list,   # 'number' | 'text'
                col_sums:   dict) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Image to Spreadsheet Export'

    thin   = Side(style='thin', color='D1D5DB')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Confidence → fill colour
    def conf_fill(conf):
        if conf < 0:    return None
        if conf >= 90:  return PatternFill('solid', fgColor='D1FAE5')  # green
        if conf >= 70:  return PatternFill('solid', fgColor='FEF3C7')  # yellow
        return             PatternFill('solid', fgColor='FEE2E2')      # red

    # Header row
    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(color='F8FAFC', bold=True, size=10, name='Calibri')

    for ci, header in enumerate(table_data[0] if table_data else [], 1):
        cell           = ws.cell(row=1, column=ci, value=header)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    # Data rows
    for ri, row in enumerate(table_data[1:], 2):
        for ci, val in enumerate(row, 1):
            col_idx = ci - 1
            ctype   = col_types[col_idx] if col_idx < len(col_types) else 'text'
            conf    = (conf_grid[ri-1][col_idx]
                       if ri-1 < len(conf_grid)
                       and col_idx < len(conf_grid[ri-1]) else -1)

            # Try numeric conversion
            num = _to_float(val) if ctype == 'number' else None
            cell = ws.cell(row=ri, column=ci,
                           value=num if num is not None else val)

            cell.border    = border
            cell.alignment = Alignment(
                horizontal='right' if ctype == 'number' else 'left',
                vertical='center'
            )
            f = conf_fill(conf)
            if f: cell.fill = f
            if ctype == 'number' and num is not None:
                cell.number_format = '#,##0.##'

    # Sum row
    n_data  = len(table_data) - 1
    sum_row = n_data + 2
    sum_fill = PatternFill('solid', fgColor='0F2942')
    sum_font = Font(color='34D399', bold=True, size=10, name='Calibri')

    for ci in range(1, len(table_data[0])+1 if table_data else 1):
        cell       = ws.cell(row=sum_row, column=ci)
        cell.fill  = sum_fill
        cell.font  = sum_font
        cell.border= border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        col_key = str(ci-1)
        if col_key in col_sums:
            cell.value = col_sums[col_key]
            cell.number_format = '#,##0.##'
        elif ci == 1:
            cell.value = 'TOTALS'

    # Column widths
    for i, col_cells in enumerate(ws.columns, 1):
        w = max((len(str(c.value)) for c in col_cells if c.value), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(w + 4, 50)

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────
# FLASK ROUTES
# ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/ocr', methods=['POST'])
def ocr_route():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    file = request.files['image']
    if not file.filename or not allowed(file.filename):
        return jsonify({'error': 'Invalid file type'}), 400

    fname    = secure_filename(f"{uuid.uuid4()}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], fname)
    try:
        file.save(filepath)
        result = run_ocr(filepath)
        if not result['data']:
            return jsonify({'error': 'No text detected. Check image quality.'}), 422
        return jsonify({'success': True, **result})
    except pytesseract.TesseractNotFoundError:
        return jsonify({'error':
            'Tesseract not found. Install: sudo apt-get install tesseract-ocr'}), 500
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if os.path.exists(filepath): os.remove(filepath)


@app.route('/export', methods=['POST'])
def export_route():
    try:
        p          = request.json or {}
        table_data = p.get('tableData', [])
        conf_grid  = p.get('confidence', [])
        num_cols   = p.get('numericCols', [])
        col_sums   = p.get('colSums', {})
        n_cols     = len(table_data[0]) if table_data else 0
        col_types  = ['number' if i in num_cols else 'text'
                      for i in range(n_cols)]
        buf  = build_excel(table_data, conf_grid, col_types, col_sums)
        return send_file(
            buf,
            mimetype=('application/vnd.openxmlformats-officedocument'
                      '.spreadsheetml.sheet'),
            as_attachment=True,
            download_name='image_to_spreadsheet_export.xlsx'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    app.run(debug=True, port=5000)
